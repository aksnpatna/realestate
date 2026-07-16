"""
etl_abs_census.py — ABS Census 2021 Demographics Pipeline
==========================================================
Downloads the ABS Census 2021 SAL (Suburb/Locality) DataPack,
parses G01 (population/age), G33 (income), G32 (tenure/ownership),
and merges into suburbs_ui_v3, overriding the OnTheHouse-sourced
demographic fields with authoritative government data.

Data source: Australian Bureau of Statistics
Licence:     Creative Commons Attribution 4.0 (CC BY 4.0)
URL:         https://www.abs.gov.au/census/find-census-data/datapacks

Legal note:  ABS data is Crown copyright, licensed for free use with
             attribution. Using this data to replace scraped demographics
             is a critical legal defensibility measure.
"""
import os
import io
import csv
import json
import re
import zipfile
import datetime
import logging
import requests
from pathlib import Path
from sqlalchemy import text
from models_v3 import SessionLocal, SuburbUIV3


def _normalise_sal_code(val):
    """Normalise the SAL_CODE_2021 cell across ABS file formats.

    The 2021 Census DataPack CSV stores SAL codes as ``SAL10001`` while the
    ASGS allocation XLSX uses bare ``10001``. We strip any leading non-digit
    characters so lookups compare consistently. Returns the string form, or
    "" if falsy.
    """
    if not val:
        return ""
    s = str(val).strip()
    return re.sub(r"^[^0-9]+", "", s)

logging.basicConfig(level=logging.INFO, format="%(asctime)s [ABS] %(message)s")
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# ABS DataPack URLs — Census 2021, SAL (Suburb/Locality) level, Short Header
# CC BY 4.0 — free, no API key required
# ---------------------------------------------------------------------------
ABS_SAL_DATAPACK_URL = (
    "https://www.abs.gov.au/census/find-census-data/datapacks/download/"
    "2021_GCP_SAL_for_AUS_short-header.zip"
)
# SAL → Postcode concordance from ABS ASGS 2021
# (URL corrected 2026-07: ABS moved the file to XLSX under /asgs-edition-3/...
# and the previous CSV path returned 404. The XLSX contains the same
# SAL_CODE_2021 / POA_CODE_2021 columns; we read either format.)
ABS_SAL_POSTCODE_URL = (
    "https://www.abs.gov.au/statistics/standards/"
    "australian-statistical-geography-standard-asgs-edition-3/"
    "jul2021-jun2026/access-and-downloads/allocation-files/SAL_2021_AUST.xlsx"
)
# Postal Areas allocation file (links Mesh Blocks -> POA / postcode). The SAL
# allocation file alone does NOT include POA, so we join MB->SAL with MB->POA
# and take the most common POA per SAL (plurality).
ABS_POA_ALLOCATION_URL = (
    "https://www.abs.gov.au/statistics/standards/"
    "australian-statistical-geography-standard-asgs-edition-3/"
    "jul2021-jun2026/access-and-downloads/allocation-files/POA_2021_AUST.xlsx"
)

CACHE_DIR = Path(__file__).parent / "data" / "abs_cache"
CACHE_DIR.mkdir(parents=True, exist_ok=True)

DATAPACK_CACHE = CACHE_DIR / "2021_GCP_SAL_AUS.zip"
CONCORDANCE_CACHE = CACHE_DIR / "SAL_2021_AUST.xlsx"
POA_ALLOCATION_CACHE = CACHE_DIR / "POA_2021_AUST.xlsx"


# ---------------------------------------------------------------------------
# Download helpers
# ---------------------------------------------------------------------------
def _download(url: str, dest: Path, label: str):
    """Download a file with progress logging. Skips if already cached."""
    if dest.exists() and dest.stat().st_size > 100_000:
        log.info(f"  Cache hit: {dest.name} ({dest.stat().st_size // 1_000_000} MB)")
        return
    log.info(f"  Downloading {label} from ABS...")
    headers = {"Accept-Encoding": "gzip, deflate", "User-Agent": "realestate-etl/3.0"}
    with requests.get(url, headers=headers, stream=True, timeout=120) as r:
        r.raise_for_status()
        total = int(r.headers.get("Content-Length", 0))
        downloaded = 0
        with open(dest, "wb") as f:
            for chunk in r.iter_content(chunk_size=1_048_576):
                f.write(chunk)
                downloaded += len(chunk)
                if total:
                    pct = downloaded * 100 // total
                    if pct % 10 == 0:
                        log.info(f"    {pct}% ({downloaded // 1_000_000} MB / {total // 1_000_000} MB)")
    log.info(f"  ✓ Downloaded {dest.name}")


# ---------------------------------------------------------------------------
# Step 1: Build SAL code → Postcode mapping
# ---------------------------------------------------------------------------
def build_sal_postcode_map() -> dict:
    """
    Build a SAL_CODE_2021 -> {"postcode","name","state"} map using the ABS
    ASGS Edition 3 allocation files. The SAL file alone has no POA column, so
    we also fetch the POA allocation file (Mesh Block -> Postal Area), then
    join MB -> SAL and MB -> POA, picking the most common POA per SAL.

    Returns dict: sal_code (str) -> {"postcode","name","state"}.
    Reads legacy CSV cache if present for backward compatibility.
    """
    _download(ABS_SAL_POSTCODE_URL, CONCORDANCE_CACHE, "SAL allocation (MB->SAL)")
    _download(ABS_POA_ALLOCATION_URL, POA_ALLOCATION_CACHE, "POA allocation (MB->POA)")
    mapping = {}

    def _add_row(sal_code, postcode, sal_name, state):
        if sal_code and postcode:
            mapping[str(sal_code).strip()] = {
                "postcode": str(postcode).strip(),
                "name": str(sal_name or "").strip().upper(),
                "state": str(state or "").strip(),
            }

    try:
        if CONCORDANCE_CACHE.suffix.lower() == ".xlsx":
            from openpyxl import load_workbook
            log.info("  Building MB->SAL & MB->POA maps (XLSX)...")

            # ----- MB -> SAL -----
            wb = load_workbook(CONCORDANCE_CACHE, read_only=True, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            header = [str(c or "").strip() for c in next(rows_iter)]

            def _idx(hlist, name):
                for i, h in enumerate(hlist):
                    if h.upper() == name.upper():
                        return i
                return None

            i_mb_sal = _idx(header, "MB_CODE_2021")
            i_sal = _idx(header, "SAL_CODE_2021")
            i_name = _idx(header, "SAL_NAME_2021")
            i_state = _idx(header, "STATE_NAME_2021")
            sal_meta = {}      # sal_code -> (name, state)
            mb_to_sal = {}     # mb_code -> sal_code
            n_rows = 0
            for row in rows_iter:
                mb = row[i_mb_sal] if i_mb_sal is not None else None
                sal = row[i_sal] if i_sal is not None else None
                if not mb or not sal:
                    continue
                mbs = str(mb).strip()
                sc = str(sal).strip()
                mb_to_sal.setdefault(mbs, sc)
                if sc not in sal_meta:
                    sal_meta[sc] = (
                        row[i_name] if i_name is not None else None,
                        row[i_state] if i_state is not None else None,
                    )
                n_rows += 1
            log.info(f"  SAL file: {len(sal_meta):,} SALs across {n_rows:,} MB rows")
            wb.close()

            # ----- MB -> POA -----
            wb = load_workbook(POA_ALLOCATION_CACHE, read_only=True, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            header = [str(c or "").strip() for c in next(rows_iter)]
            i_mb_poa = _idx(header, "MB_CODE_2021")
            i_poa = _idx(header, "POA_CODE_2021")
            if i_mb_poa is None or i_poa is None:
                log.error("  POA XLSX missing MB_CODE_2021 / POA_CODE_2021 columns")
                return mapping
            # sal_code -> {poa_code: count}  for plurality vote
            sal_poa_tally = {}
            n_rows = 0
            for row in rows_iter:
                mb = row[i_mb_poa]
                poa = row[i_poa]
                if not mb or not poa:
                    continue
                sc = mb_to_sal.get(str(mb).strip())
                if not sc:
                    continue
                sal_poa_tally.setdefault(sc, {})
                pc = str(poa).strip()
                sal_poa_tally[sc][pc] = sal_poa_tally[sc].get(pc, 0) + 1
                n_rows += 1
            log.info(f"  POA file: matched {len(sal_poa_tally):,} SALs across {n_rows:,} MB rows")
            wb.close()

            # ----- SAL -> plurality POA -----
            for sc, tally in sal_poa_tally.items():
                poa = max(tally, key=tally.get)
                name, state = sal_meta.get(sc, (None, None))
                _add_row(sc, poa, name, state)
        else:
            # Legacy CSV cache
            with open(CONCORDANCE_CACHE, newline="", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    sal_code = row.get("SAL_CODE_2021", "").strip()
                    postcode = row.get("POA_CODE_2021", row.get("POSTCODE_2021", "")).strip()
                    sal_name = row.get("SAL_NAME_2021", "").strip().upper()
                    state = row.get("STATE_NAME_2021", "").strip()
                    _add_row(sal_code, postcode, sal_name, state)
    except Exception as e:
        log.error(f"  Failed to parse concordance: {e}")
    log.info(f"  Built SAL→Postcode map: {len(mapping):,} entries")
    return mapping


# ---------------------------------------------------------------------------
# Step 2: Parse Census tables from ZIP
# ---------------------------------------------------------------------------
def parse_census_tables(sal_codes: set) -> dict:
    """
    Opens the ABS DataPack ZIP in-memory and parses:
      - G01: Total persons, age summary
      - G33: Income bands (weekly personal income)
      - G32: Tenure type (owner, renter, etc.)
    Returns dict: sal_code → parsed demographics dict
    """
    _download(ABS_SAL_DATAPACK_URL, DATAPACK_CACHE, "Census 2021 SAL DataPack (102 MB)")

    results = {}

    log.info("  Parsing Census tables from DataPack ZIP...")
    with zipfile.ZipFile(DATAPACK_CACHE, "r") as zf:
        all_files = zf.namelist()

        # Identify table files — short header format names like:
        # 2021Census_G01_AUS_SAL.csv, 2021Census_G33_AUS_SAL.csv, etc.
        def find_table(table_id):
            matches = [f for f in all_files if f"_{table_id}_" in f and "SAL" in f and f.endswith(".csv")]
            return matches[0] if matches else None

        # ── G01: Selected Person Characteristics ──────────────────────────
        g01_file = find_table("G01")
        if g01_file:
            log.info(f"  Parsing {g01_file}...")
            with zf.open(g01_file) as f:
                reader = csv.DictReader(io.TextIOWrapper(f, encoding="utf-8-sig"))
                for row in reader:
                    sal = _normalise_sal_code(row.get("SAL_CODE_2021", ""))
                    if sal not in sal_codes:
                        continue
                    if sal not in results:
                        results[sal] = {}
                    try:
                        results[sal]["population_2021"] = int(float(row.get("Tot_P_P", 0) or 0))
                        # G01 has no median age — that lives in G02 (below).
                    except Exception:
                        pass

        # ── G02: Selected Medians and Averages ──────────────────────────
        g02_file = find_table("G02")
        if g02_file:
            log.info(f"  Parsing {g02_file}...")
            with zf.open(g02_file) as f:
                reader = csv.DictReader(io.TextIOWrapper(f, encoding="utf-8-sig"))
                for row in reader:
                    sal = _normalise_sal_code(row.get("SAL_CODE_2021", ""))
                    if sal not in sal_codes:
                        continue
                    if sal not in results:
                        results[sal] = {}
                    try:
                        results[sal]["median_age_persons"] = _safe_int(row.get("Median_age_persons"))
                        results[sal]["median_rent_weekly_abs"] = _safe_float(row.get("Median_rent_weekly"))
                        results[sal]["median_mortgage_monthly_abs"] = _safe_float(row.get("Median_mortgage_repay_monthly"))
                        results[sal]["median_tot_hhd_inc_weekly_abs"] = _safe_float(row.get("Median_tot_hhd_inc_weekly"))
                        results[sal]["median_tot_prsnl_inc_weekly_abs"] = _safe_float(row.get("Median_tot_prsnl_inc_weekly"))
                        hh = _safe_float(row.get("Average_household_size"))
                        if hh is not None:
                            results[sal]["average_household_size"] = hh
                    except Exception:
                        pass


        # ── G01 Age Bands (already parsed above — extend with distribution) ─
        # G01 has Age_0_4_yr_P, Age_5_14_yr_P, ..., Age_85ov_P which are
        # the summarised five-year/ten-year age bands. We prefer these over the
        # single-year G04A/G04B split because they match what the UI will display.
        g01_age_file = find_table("G01")
        if g01_age_file:
            age_bands = [
                ("Age_0_4_yr_P", "0-4"),
                ("Age_5_14_yr_P", "5-14"),
                ("Age_15_19_yr_P", "15-19"),
                ("Age_20_24_yr_P", "20-24"),
                ("Age_25_34_yr_P", "25-34"),
                ("Age_35_44_yr_P", "35-44"),
                ("Age_45_54_yr_P", "45-54"),
                ("Age_55_64_yr_P", "55-64"),
                ("Age_65_74_yr_P", "65-74"),
                ("Age_75_84_yr_P", "75-84"),
                ("Age_85ov_P", "85+"),
            ]
            with zf.open(g01_age_file) as f:
                reader = csv.DictReader(io.TextIOWrapper(f, encoding="utf-8-sig"))
                for row in reader:
                    sal = _normalise_sal_code(row.get("SAL_CODE_2021", ""))
                    if sal not in sal_codes:
                        continue
                    if sal not in results:
                        results[sal] = {}
                    dist = {}
                    for col, label in age_bands:
                        val = _safe_float(row.get(col))
                        if val is not None:
                            dist[label] = val
                    if dist:
                        total = sum(dist.values()) or 1
                        dist_pct = {k: round(v / total * 100, 1) for k, v in dist.items()}
                        results[sal]["age_distribution"] = dist_pct
                        dominant = max(dist, key=dist.get)
                        results[sal]["predominant_age_group"] = dominant

        # ── G33: Household Income — SAL DataPack uses HI_ prefix + _Tot suffix
        g33_file = find_table("G33")
        if g33_file:
            log.info(f"  Parsing {g33_file}...")
            income_map = {
                "Negative_Nil_income_Tot": "Nil/Negative",
                "HI_1_149_Tot":              "$1-$149",
                "HI_150_299_Tot":            "$150-$299",
                "HI_300_399_Tot":            "$300-$399",
                "HI_400_499_Tot":            "$400-$499",
                "HI_500_649_Tot":            "$500-$649",
                "HI_650_799_Tot":            "$650-$799",
                "HI_800_999_Tot":            "$800-$999",
                "HI_1000_1249_Tot":          "$1,000-$1,249",
                "HI_1250_1499_Tot":          "$1,250-$1,499",
                "HI_1500_1749_Tot":          "$1,500-$1,749",
                "HI_1750_1999_Tot":          "$1,750-$1,999",
                "HI_2000_2499_Tot":          "$2,000-$2,499",
                "HI_2500_2999_Tot":          "$2,500-$2,999",
                "HI_3000_3499_Tot":          "$3,000-$3,499",
                "HI_3500_3999_Tot":          "$3,500-$3,999",
                "HI_4000_more_Tot":          "$4,000+",
            }
            with zf.open(g33_file) as f:
                reader = csv.DictReader(io.TextIOWrapper(f, encoding="utf-8-sig"))
                for row in reader:
                    sal = _normalise_sal_code(row.get("SAL_CODE_2021", ""))
                    if sal not in sal_codes:
                        continue
                    if sal not in results:
                        results[sal] = {}
                    dist = {}
                    for col, label in income_map.items():
                        val = _safe_float(row.get(col))
                        if val is not None:
                            dist[label] = val
                    if dist:
                        total = sum(dist.values()) or 1
                        dist_pct = {k: round(v / total * 100, 1) for k, v in dist.items()}
                        results[sal]["income_distribution"] = dist_pct
                        dominant = max(dist_pct, key=dist_pct.get)
                        results[sal]["predominant_income_band"] = dominant

        # ── G37: Tenure and Landlord Type by Dwelling Structure ──────────
        # (Note: in the 2021 Census SAL DataPack, tenure lives in G37, NOT G32.
        # G32 covers family income and has no tenure columns despite the
        # previous version of this script attempting to read them from there.)
        g37_file = find_table("G37")
        if g37_file:
            log.info(f"  Parsing {g37_file}...")
            with zf.open(g37_file) as f:
                reader = csv.DictReader(io.TextIOWrapper(f, encoding="utf-8-sig"))
                for row in reader:
                    sal = _normalise_sal_code(row.get("SAL_CODE_2021", ""))
                    if sal not in sal_codes:
                        continue
                    if sal not in results:
                        results[sal] = {}
                    own_outright = _safe_float(row.get("O_OR_Total"))   # Owned outright
                    own_mortgage = _safe_float(row.get("O_MTG_Total"))  # Owned with a mortgage
                    # Total renting = real-estate + state housing authority + community
                    # housing provider + person-not-in-same-hh + other landlord type.
                    rent_rea    = _safe_float(row.get("R_RE_Agt_Total"))         # real-estate agent
                    rent_state  = _safe_float(row.get("R_ST_h_auth_Total"))     # state housing authority
                    rent_comm   = _safe_float(row.get("R_Com_Hp_Total"))       # community housing provider
                    rent_person = _safe_float(row.get("R_Psn_not_in_s_hh_Total"))
                    rent_other  = _safe_float(row.get("R_Ot_landld_typ_Total"))
                    total_hh    = _safe_float(row.get("Total_Total"))
                    if total_hh and total_hh > 0:
                        owner_pct = round(((own_mortgage or 0) + (own_outright or 0)) / total_hh * 100, 1)
                        renting = ((rent_rea or 0) + (rent_state or 0) + (rent_comm or 0)
                                  + (rent_person or 0) + (rent_other or 0))
                        investor_pct = round(renting / total_hh * 100, 1)
                        results[sal]["owner_occupier_rate"] = owner_pct
                        results[sal]["investor_rate"] = investor_pct

    log.info(f"  ✓ Parsed {len(results):,} SAL records from Census")
    return results


def _safe_float(val):
    try:
        return float(val) if val not in (None, "", "...", "-") else None
    except (ValueError, TypeError):
        return None


def _safe_int(val):
    try:
        return int(float(val)) if val not in (None, "", "...", "-") else None
    except (ValueError, TypeError):
        return None


# ---------------------------------------------------------------------------
# Step 3: Match census data to our suburbs and update DB
# ---------------------------------------------------------------------------
def run_abs_integration(batch_size: int = 500,
                        state: str | None = None,
                        limit: int | None = None,
                        dry_run: bool = False,
                        sample_ids: list | None = None):
    """
    Main entry point.
    Downloads ABS data, matches to suburbs_ui_v3 by postcode+name,
    and overwrites demographic fields with authoritative ABS data.
    Sets abs_demographics_sourced = True on every updated row.

    Args:
        batch_size:  commit batch size (default 500)
        state:       optional 2-3 letter state filter ("VIC", "NSW", ...) — None = all
        limit:       process at most `limit` suburbs — useful for spot-checks
        dry_run:     if True, compute ABS-vs-OTH diff and print it but DO NOT write
                     to the DB. abs_demographics_sourced stays False.
        sample_ids:  optional explicit list of suburb IDs to scope strictly (useful
                     for spot-checks on known suburbs). If non-empty, this wins
                     over state/limit.
    """
    log.info("=" * 60)
    log.info("ABS Census 2021 Demographics Pipeline — Starting")
    log.info(f"  state={state or 'ALL'}  limit={limit or 'none'}  "
             f"dry_run={dry_run}  sample_ids={len(sample_ids) if sample_ids else 0}")
    log.info("=" * 60)

    # 1. Build postcode concordance
    sal_map = build_sal_postcode_map()
    if not sal_map:
        log.error("Could not build SAL map — aborting.")
        return

    # 2. Build reverse lookup: (UPPER_NAME, POSTCODE) → SAL_CODE
    name_postcode_to_sal: dict[tuple, str] = {}
    postcode_to_sals: dict[str, list] = {}
    for sal_code, info in sal_map.items():
        key = (info["name"], info["postcode"])
        name_postcode_to_sal[key] = sal_code
        postcode_to_sals.setdefault(info["postcode"], []).append(sal_code)

    # 3. Load suburb list from DB
    db = SessionLocal()
    try:
        q = db.query(
            SuburbUIV3.id, SuburbUIV3.name, SuburbUIV3.postcode,
            SuburbUIV3.population_2021, SuburbUIV3.median_age,
            SuburbUIV3.owner_occupier_rate, SuburbUIV3.investor_rate,
            SuburbUIV3.predominant_age_group,
        ).filter(SuburbUIV3.is_enriched == True)
        if sample_ids:
            q = q.filter(SuburbUIV3.id.in_(list(sample_ids)))
        if state:
            q = q.filter(SuburbUIV3.state == state.upper())
        if limit:
            q = q.limit(limit)
        suburbs = q.all()
    finally:
        db.close()

    log.info(f"  Found {len(suburbs):,} enriched suburbs to match against ABS")

    # 4. Match each suburb to a SAL code
    matched_sal_codes: dict[str, str] = {}  # suburb_id → SAL code
    unmatched = 0
    for sid, name, postcode, *_rest in suburbs:
        upper_name = name.upper().strip() if name else ""
        # Try exact name+postcode match first
        sal = name_postcode_to_sal.get((upper_name, postcode or ""))
        if not sal:
            # Fall back to postcode-only match (take first SAL in that postcode)
            candidates = postcode_to_sals.get(postcode or "", [])
            sal = candidates[0] if len(candidates) == 1 else None
            if not sal:
                unmatched += 1
                continue
        matched_sal_codes[sid] = sal

    log.info(f"  Matched {len(matched_sal_codes):,} suburbs, {unmatched} unmatched")

    # 5. Parse only the SAL codes we actually need (saves memory)
    needed_sals = set(matched_sal_codes.values())
    census_data = parse_census_tables(needed_sals)

    # ── Dry-run: print OTH-vs-ABS diff and exit without writing ────────────
    if dry_run:
        log.info("=" * 60)
        log.info("DRY RUN — printing OTH → ABS diff, no DB writes")
        log.info("=" * 60)
        # Keep a summarised row info keyed by suburb id
        oth_by_id = {row[0]: row for row in suburbs}
        print(f"\n{'suburb_id':<28}{'pop_OTH':>8}{'pop_ABS':>9}{'Δ%':>6}  "
              f"{'age_OTH':>7}{'age_ABS':>8}  {'own%_OTH':>8}{'own%_ABS':>9}")
        print("-" * 92)
        for sid, sal in matched_sal_codes.items():
            c = census_data.get(sal, {})
            if not c:
                continue
            _, name, postcode, pop_oth, age_oth, own_oth, inv_oth, dom_oth = oth_by_id[sid]
            pop_abs = c.get("population_2021")
            age_abs = c.get("median_age_persons")
            own_abs = c.get("owner_occupier_rate")
            pop_diff = (f"{(pop_oth - pop_abs) / pop_abs * 100:+.0f}%" if (pop_oth and pop_abs) else "—")
            own_oth_s = f"{own_oth:.0f}" if own_oth is not None else "—"
            own_abs_s = f"{own_abs:.1f}" if own_abs is not None else "—"
            print(f"{sid:<28}{pop_oth or 0:>8}{(pop_abs or 0):>9}{pop_diff:>6}  "
                  f"{age_oth or 0:>7}{(age_abs or 0):>8}  {own_oth_s:>8}{own_abs_s:>9}")
        log.info("=" * 60)
        log.info(f"DRY RUN complete — 0 DB writes (would update {len(matched_sal_codes):,} suburbs if not dry)")
        log.info("=" * 60)
        return

    # 6. Update DB in batches
    db = SessionLocal()
    updated = 0
    skipped = 0
    now = datetime.datetime.utcnow()

    try:
        suburb_iter = list(matched_sal_codes.items())
        for i in range(0, len(suburb_iter), batch_size):
            batch = suburb_iter[i:i + batch_size]
            for sid, sal in batch:
                c = census_data.get(sal)
                if not c:
                    skipped += 1
                    continue

                sourced_fields = []
                update_vals = {
                    "abs_demographics_sourced": True,
                    "abs_etl_run_date": now,
                }

                def _set(field, census_key):
                    val = c.get(census_key)
                    if val is not None:
                        update_vals[field] = val
                        sourced_fields.append(field)

                _set("population_2021", "population_2021")
                _set("median_age", "median_age_persons")
                _set("owner_occupier_rate", "owner_occupier_rate")
                _set("investor_rate", "investor_rate")
                _set("predominant_age_group", "predominant_age_group")
                _set("average_household_size", "average_household_size")

                # Update demographics_detail JSON with ABS data merged in
                demo_detail = {}
                if c.get("age_distribution"):
                    demo_detail["age_distribution"] = c["age_distribution"]
                    sourced_fields.append("age_distribution")
                if c.get("income_distribution"):
                    demo_detail["income_distribution"] = c["income_distribution"]
                    sourced_fields.append("income_distribution")
                if c.get("predominant_income_band"):
                    demo_detail["predominant_income_band"] = c["predominant_income_band"]
                # median_annual_income_abs derived from G02 household weekly median
                weekly = c.get("median_tot_prsnl_inc_weekly_abs")
                if weekly:
                    demo_detail["median_weekly_income_abs"] = weekly
                    demo_detail["median_annual_income_abs"] = round(weekly * 52)
                    sourced_fields.append("median_annual_income")
                if c.get("median_rent_weekly_abs"):
                    demo_detail["median_rent_weekly_abs"] = c["median_rent_weekly_abs"]
                    sourced_fields.append("median_rent_weekly_abs")
                if c.get("median_mortgage_monthly_abs"):
                    demo_detail["median_mortgage_monthly_abs"] = c["median_mortgage_monthly_abs"]
                    sourced_fields.append("median_mortgage_monthly_abs")
                if c.get("median_tot_hhd_inc_weekly_abs"):
                    demo_detail["median_hhd_inc_weekly_abs"] = c["median_tot_hhd_inc_weekly_abs"]
                    sourced_fields.append("median_hhd_inc_weekly_abs")

                if demo_detail:
                    # Merge ABS data UNDER existing demographics_detail so OTH keys
                    # (area_sqkm, parks_count, predominant_occupation, etc.) survive.
                    existing = db.query(SuburbUIV3.demographics_detail).filter(
                        SuburbUIV3.id == sid
                    ).scalar()
                    merged = {**(existing or {}), **demo_detail}
                    update_vals["demographics_detail"] = merged

                update_vals["abs_sourced_fields"] = sourced_fields

                db.query(SuburbUIV3).filter(SuburbUIV3.id == sid).update(
                    update_vals, synchronize_session=False
                )
                updated += 1

            db.commit()
            log.info(f"  Batch {i // batch_size + 1}: committed {min(i + batch_size, len(suburb_iter))} / {len(suburb_iter)}")

    except Exception as e:
        log.error(f"Pipeline error: {e}")
        db.rollback()
        raise
    finally:
        db.close()

    log.info("=" * 60)
    log.info(f"ABS Census Pipeline Complete")
    log.info(f"  Updated:  {updated:,} suburbs with real ABS 2021 data")
    log.info(f"  Skipped:  {skipped:,} (no census data for SAL code)")
    log.info(f"  Unmatched:{unmatched:,} (no SAL code found for suburb)")
    log.info(f"  Source:   ABS Census 2021 — CC BY 4.0")
    log.info("=" * 60)


if __name__ == "__main__":
    import argparse
    p = argparse.ArgumentParser(description="ABS Census 2021 demographics enrichment")
    p.add_argument("--state", type=str, default=None, help="State filter e.g. VIC")
    p.add_argument("--limit", type=int, default=None, help="Limit number of suburbs processed")
    p.add_argument("--batch", type=int, default=500, help="Commit batch size")
    p.add_argument("--dry-run", action="store_true",
                   help="Compute ABS-vs-OTH diff and print; do NOT write to DB")
    p.add_argument("--sample", type=str, default=None,
                   help="Comma-separated suburb IDs to scope strictly")
    args = p.parse_args()
    sample_ids = [s.strip() for s in args.sample.split(",")] if args.sample else None
    run_abs_integration(batch_size=args.batch, state=args.state, limit=args.limit,
                        dry_run=args.dry_run, sample_ids=sample_ids)
