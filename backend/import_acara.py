"""
ACARA School ICSEA Import Script
Merges School Profile 2025 (ICSEA scores) + School Location 2025 (lat/lon, suburb)
into the suburbs_all database, enriching the school quality signals.
"""
import os
import json
import sys
import openpyxl
from collections import defaultdict
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker

DATABASE_URL = os.getenv("DATABASE_URL", "postgresql://realestate_user:realestate_pass@db:5432/realestate")
engine = create_engine(DATABASE_URL)
SessionLocal = sessionmaker(bind=engine)

TARGET_STATES = {'VIC', 'NSW', 'QLD', 'TAS', 'SA'}

PROFILE_FILE = "/app/SchoolProfile2025.xlsx"
LOCATION_FILE = "/app/SchoolLocation2025.xlsx"
UI_DATA_FILE = "/app/suburbs_data.json"


def load_location_data():
    """Load school lat/lon keyed by ACARA SML ID."""
    print("Loading school locations...")
    wb = openpyxl.load_workbook(LOCATION_FILE, read_only=True)
    ws = wb['SchoolLocations 2025']
    headers = None
    locations = {}
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = {v: i for i, v in enumerate(row)}
            continue
        acara_id = row[headers['ACARA SML ID']]
        lat = row[headers['Latitude']]
        lon = row[headers['Longitude']]
        suburb = row[headers['Suburb']]
        state = row[headers['State']]
        postcode = str(row[headers['Postcode']] or '').zfill(4)
        if state in TARGET_STATES and lat and lon:
            locations[acara_id] = {
                'lat': lat, 'lon': lon,
                'suburb': suburb, 'state': state, 'postcode': postcode
            }
    print(f"  Loaded {len(locations)} school locations in target states.")
    return locations


def load_profile_data(locations):
    """Load school profiles with ICSEA, merging location data."""
    print("Loading school profiles (ICSEA)...")
    wb = openpyxl.load_workbook(PROFILE_FILE, read_only=True)
    ws = wb['SchoolProfile 2025']
    headers = None

    # Group schools by suburb+state+postcode
    schools_by_suburb = defaultdict(list)
    
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = {v: i for i, v in enumerate(row)}
            continue

        acara_id = row[headers['ACARA SML ID']]
        loc = locations.get(acara_id)
        if not loc:
            continue  # Skip schools without location data
        
        icsea = row[headers['ICSEA']]
        icsea_percentile = row[headers['ICSEA Percentile']]
        school_name = row[headers['School Name']]
        school_type = row[headers.get('School Type', -1)] if 'School Type' in headers else None
        year_range = row[headers.get('Year Range', -1)] if 'Year Range' in headers else None
        
        if not icsea:
            continue  # No ICSEA score, skip
        
        suburb_key = (loc['suburb'].title(), loc['state'].upper(), loc['postcode'])
        
        schools_by_suburb[suburb_key].append({
            'name': school_name,
            'icsea': int(icsea),
            'icseaPercentile': int(icsea_percentile) if icsea_percentile else None,
            'type': school_type or 'Unknown',
            'yearRange': year_range,
            'lat': loc['lat'],
            'lon': loc['lon'],
        })
    
    print(f"  Found schools in {len(schools_by_suburb)} unique suburbs.")
    return schools_by_suburb


def compute_school_quality_score(schools):
    """
    Compute a 0-10 school quality score from ICSEA.
    ICSEA mean is 1000, std ~100. So:
      - 900  = bottom ~16%  → score ~3
      - 1000 = median       → score ~5
      - 1100 = top ~16%     → score ~7
      - 1200+= elite        → score ~9-10
    """
    if not schools:
        return None
    avg_icsea = sum(s['icsea'] for s in schools) / len(schools)
    # Normalize: ICSEA 800=1, 1000=5, 1200=9, cap at 10
    score = max(1.0, min(10.0, (avg_icsea - 800) / 50))
    return round(score, 1)


def update_db_and_json(schools_by_suburb):
    """Update SuburbUIV3 DB with ACARA school data."""
    from models_v3 import SuburbUIV3
    db = SessionLocal()
    updated_db = 0
    
    print("\nUpdating database SuburbUIV3 with ACARA school data...")
    for (suburb_name, state, postcode), schools in schools_by_suburb.items():
        v3 = db.query(SuburbUIV3).filter(
            SuburbUIV3.name.ilike(suburb_name),
            SuburbUIV3.state.ilike(state),
            SuburbUIV3.postcode == postcode
        ).first()
        
        if v3:
            quality_score = compute_school_quality_score(schools)
            v3.avg_icsea = round(sum(s['icsea'] for s in schools) / len(schools), 0)
            v3.school_count = len(schools)
            v3.school_quality = quality_score
            updated_db += 1
    
    db.commit()
    db.close()
    print(f"  Updated {updated_db} suburbs in database.")
    
    # We no longer update the legacy suburbs_data.json file since V3 uses the database exclusively.
    # Keep the print statement to match expectations.
    print(f"  Skipped updating legacy suburbs_data.json (UI data).")


def main():
    print("=== ACARA School ICSEA Import ===\n")
    
    locations = load_location_data()
    schools_by_suburb = load_profile_data(locations)
    update_db_and_json(schools_by_suburb)
    
    print("\n✅ Import complete!")


if __name__ == "__main__":
    main()
